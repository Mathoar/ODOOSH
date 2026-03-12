import base64
import io
import logging

from odoo import models, fields, api, _
from odoo.exceptions import UserError

_logger = logging.getLogger(__name__)

try:
    from openpyxl import load_workbook
except ImportError:
    _logger.warning("openpyxl not installed. Excel import will not work.")

GROUPE_FOURNISSEUR_MAPPING = {
    'ASIE-INDE': 'asie_inde',
    'AISIE-INDE': 'asie_inde',
    'ASIE -INDE': 'asie_inde',
    'EUROPE': 'europe',
    'LOCAL': 'local',
    'AFRIQUE': 'afrique',
    'AMERIQUE': 'amerique',
    'OCEANIE': 'oceanie',
}

GROUPE_CLIENT_MAPPING = {
    'GBH': 'gbh',
    'SMDIS': 'smdis',
    'CELLULE U': 'cellule_u',
    'CAILLE': 'caille',
    'DISTRIMASCAREIGNES': 'distrimascareignes',
    'IBL': 'ibl',
    'TK': 'tk',
    'LOT': 'lot',
    'GMS': 'gms',
}

ENSEIGNE_MAPPING = {
    'AUCHAN': 'auchan',
    'CARREFOUR': 'carrefour',
    'CARREFOUR MARKET': 'carrefour_market',
    'CARREFOUR CITY': 'carrefour_city',
    'PROMOCASH': 'promocash',
    'U': 'u',
    'LECLERC': 'leclerc',
    'INTERMARK': 'intermark',
    'RUN MARKET': 'run_market',
    'COCCI MARKET': 'cocci_market',
    'LEADER PRICE': 'leader_price',
    'GEL OI': 'gel_oi',
    'INDEPENDANT': 'independant',
}

CATEGORIE_CLIENT_MAPPING = {
    'GMS': 'GMS',
    'PLATE FORME': 'PLATE-FORME',
    'PLATE-FORME': 'PLATE-FORME',
    'GROSSISTE': 'GROSSISTE',
    'CHR': 'CHR',
    'DIVERS': 'DIVERS',
}

CONDITION_PAIEMENT_MAPPING = {
    '00': 'Comptant',
    '01': '15 jours',
    '02': '30 jours',
    '03': '60 jours',
}

TVA_MAPPING = {
    2.1: 'TVA 2.1% DOM',
    8.5: 'TVA 8.5% DOM',
    0.0: 'Exonéré',
}


class FraispeiImportWizard(models.TransientModel):
    _name = 'fraispei.import.wizard'
    _description = 'Assistant d\'import FRAIS PEI'

    import_type = fields.Selection([
        ('suppliers', 'Fournisseurs'),
        ('clients', 'Clients'),
        ('products', 'Produits'),
    ], string='Type d\'import', required=True, default='suppliers')

    file_data = fields.Binary(string='Fichier Excel', required=True)
    file_name = fields.Char(string='Nom du fichier')
    dry_run = fields.Boolean(string='Mode simulation', default=True,
                             help='Si coché, aucune modification ne sera effectuée')
    import_limit = fields.Integer(string='Limite d\'import', default=0,
                                  help='0 = tout importer')

    result_text = fields.Text(string='Résultat', readonly=True)
    state = fields.Selection([
        ('draft', 'Configuration'),
        ('done', 'Terminé'),
    ], default='draft')

    def _get_cell(self, row, idx):
        if idx is None or idx >= len(row):
            return None
        val = row[idx]
        if val is None:
            return None
        if isinstance(val, str):
            return val.strip() or None
        return val

    def _get_float(self, row, idx, default=0.0):
        val = self._get_cell(row, idx)
        if val is None:
            return default
        try:
            return float(val)
        except (ValueError, TypeError):
            return default

    def _get_int(self, row, idx, default=0):
        val = self._get_cell(row, idx)
        if val is None:
            return default
        try:
            return int(val)
        except (ValueError, TypeError):
            return default

    def _load_workbook(self):
        if not self.file_data:
            raise UserError(_("Veuillez sélectionner un fichier Excel."))
        data = base64.b64decode(self.file_data)
        return load_workbook(io.BytesIO(data), read_only=True)

    def _find_or_create_partner_category(self, name):
        Category = self.env['res.partner.category']
        cat = Category.search([('name', '=', name)], limit=1)
        if cat:
            return cat.id
        return Category.create({'name': name}).id

    def _find_or_create_product_category(self, name, parent_id=None):
        ProductCategory = self.env['product.category']
        domain = [('name', '=', name)]
        if parent_id:
            domain.append(('parent_id', '=', parent_id))
        cat = ProductCategory.search(domain, limit=1)
        if cat:
            return cat.id
        vals = {'name': name}
        if parent_id:
            vals['parent_id'] = parent_id
        return ProductCategory.create(vals).id

    # =========================================================================
    # IMPORT FOURNISSEURS
    # =========================================================================
    def _import_suppliers(self):
        wb = self._load_workbook()
        sheet_name = 'TRAME FRS'
        if sheet_name not in wb.sheetnames:
            raise UserError(_(f"Feuille '{sheet_name}' non trouvée dans le fichier."))

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:]

        if self.import_limit > 0:
            data_rows = data_rows[:self.import_limit]

        Partner = self.env['res.partner']
        Country = self.env['res.country']
        stats = {'total': len(data_rows), 'created': 0, 'updated': 0, 'skipped': 0, 'errors': 0}
        country_cache = {}

        for i, row in enumerate(data_rows):
            try:
                nom = self._get_cell(row, 2)
                if not nom:
                    stats['skipped'] += 1
                    continue

                code_frs = self._get_cell(row, 0)
                values = {
                    'name': nom,
                    'is_company': True,
                    'supplier_rank': 1,
                    'customer_rank': 0,
                    'ref': str(code_frs) if code_frs else False,
                }

                raison_sociale = self._get_cell(row, 3)
                if raison_sociale and raison_sociale != nom:
                    values['name'] = raison_sociale
                    values['x_nom_commercial'] = nom

                adresse = self._get_cell(row, 4)
                if adresse:
                    lines = str(adresse).strip().split('\n')
                    values['street'] = lines[0]
                    if len(lines) > 1:
                        values['street2'] = '\n'.join(lines[1:])

                cp = self._get_cell(row, 5)
                if cp:
                    values['zip'] = str(cp)
                ville = self._get_cell(row, 6)
                if ville:
                    values['city'] = ville

                is_france = self._get_float(row, 12)
                groupe_frs = self._get_cell(row, 15)
                country_name = None
                if is_france == 1:
                    country_name = 'France'
                elif groupe_frs:
                    g = groupe_frs.upper()
                    if 'INDE' in g:
                        country_name = 'India'
                    elif 'EUROPE' in g or 'LOCAL' in g:
                        country_name = 'France'

                if country_name:
                    if country_name not in country_cache:
                        c = Country.search([('name', 'ilike', country_name)], limit=1)
                        country_cache[country_name] = c.id if c else False
                    if country_cache[country_name]:
                        values['country_id'] = country_cache[country_name]

                phone = self._get_cell(row, 9)
                if phone:
                    values['phone'] = str(phone)
                email = self._get_cell(row, 13)
                if email:
                    values['email'] = email
                website = self._get_cell(row, 14)
                if website:
                    values['website'] = website

                grp = GROUPE_FOURNISSEUR_MAPPING.get(
                    (groupe_frs or '').upper().strip(), 'autres')
                values['x_groupe_fournisseur'] = grp

                x_code = self._get_cell(row, 1)
                if x_code:
                    values['x_code_interne'] = str(x_code)
                x_aux = self._get_cell(row, 19)
                if x_aux:
                    values['x_compte_auxiliaire'] = str(x_aux)
                if self._get_float(row, 21) == 1:
                    values['x_is_transitaire'] = True
                dlai = self._get_int(row, 22)
                if dlai:
                    values['x_delai_livraison'] = dlai

                commentaires = self._get_cell(row, 17)
                if commentaires:
                    values['comment'] = commentaires

                if not self.dry_run:
                    domain = ([('ref', '=', values['ref'])] if values.get('ref')
                              else [('name', '=', values['name']),
                                    ('supplier_rank', '>', 0)])
                    existing = Partner.search(domain, limit=1)
                    if existing:
                        existing.write(values)
                        stats['updated'] += 1
                    else:
                        Partner.create(values)
                        stats['created'] += 1
                else:
                    stats['created'] += 1

            except Exception as e:
                stats['errors'] += 1
                _logger.error("Import fournisseur ligne %s: %s", i + 2, e)

        wb.close()
        return stats

    # =========================================================================
    # IMPORT CLIENTS
    # =========================================================================
    def _import_clients(self):
        wb = self._load_workbook()
        sheet_name = 'trame client'
        if sheet_name not in wb.sheetnames:
            raise UserError(_(f"Feuille '{sheet_name}' non trouvée dans le fichier."))

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:]

        if self.import_limit > 0:
            data_rows = data_rows[:self.import_limit]

        Partner = self.env['res.partner']
        stats = {'total': len(data_rows), 'created': 0, 'updated': 0, 'skipped': 0, 'errors': 0}
        category_cache = {}
        pricelist_cache = {}
        payment_cache = {}

        for i, row in enumerate(data_rows):
            try:
                nom = self._get_cell(row, 1)
                if not nom:
                    stats['skipped'] += 1
                    continue

                code = self._get_cell(row, 0)
                values = {
                    'name': nom,
                    'is_company': True,
                    'customer_rank': 1,
                    'ref': str(code) if code else False,
                }

                phone = self._get_cell(row, 3)
                if phone:
                    values['phone'] = str(phone)
                gsm = self._get_cell(row, 6)
                if gsm:
                    values['mobile'] = str(gsm)
                email = self._get_cell(row, 31)
                if email:
                    values['email'] = email
                siret = self._get_cell(row, 42)
                if siret:
                    values['siret'] = str(siret).replace(' ', '')

                plafond = self._get_cell(row, 18)
                if plafond:
                    try:
                        values['credit_limit'] = float(plafond)
                    except (ValueError, TypeError):
                        pass

                cat_name = self._get_cell(row, 16)
                if cat_name:
                    norm = CATEGORIE_CLIENT_MAPPING.get(cat_name.upper(), cat_name)
                    if norm not in category_cache:
                        category_cache[norm] = self._find_or_create_partner_category(norm)
                    values['category_id'] = [(6, 0, [category_cache[norm]])]

                code_tarif = self._get_cell(row, 22)
                if code_tarif:
                    tarif_map = {'00': 'Tarif Base', '01': 'Tarif T1', '02': 'Tarif T2',
                                 '03': 'Tarif T3', '04': 'Tarif T4', '05': 'Tarif T5',
                                 '06': 'Tarif T6'}
                    t_name = tarif_map.get(str(code_tarif).zfill(2), 'Tarif Base')
                    if t_name not in pricelist_cache:
                        pl = self.env['product.pricelist'].search(
                            [('name', 'ilike', t_name)], limit=1)
                        pricelist_cache[t_name] = pl.id if pl else False
                    if pricelist_cache[t_name]:
                        values['property_product_pricelist'] = pricelist_cache[t_name]

                code_reg = self._get_cell(row, 21)
                if code_reg:
                    term_name = CONDITION_PAIEMENT_MAPPING.get(str(code_reg).zfill(2))
                    if term_name:
                        if term_name not in payment_cache:
                            pt = self.env['account.payment.term'].search(
                                [('name', 'ilike', term_name)], limit=1)
                            payment_cache[term_name] = pt.id if pt else False
                        if payment_cache[term_name]:
                            values['property_payment_term_id'] = payment_cache[term_name]

                nom_gr = self._get_cell(row, 12)
                if nom_gr:
                    values['x_groupe_client'] = GROUPE_CLIENT_MAPPING.get(
                        nom_gr.upper(), 'autres')
                nom_ens = self._get_cell(row, 14)
                if nom_ens:
                    values['x_enseigne'] = ENSEIGNE_MAPPING.get(
                        nom_ens.upper(), 'independant')

                tx_rem = self._get_cell(row, 33)
                if tx_rem:
                    values['x_taux_remise'] = float(tx_rem) if tx_rem else 0
                tx_rfa = self._get_cell(row, 34)
                if tx_rfa:
                    values['x_taux_rfa'] = float(tx_rfa) if tx_rfa else 0

                num_compta = self._get_cell(row, 39)
                if num_compta:
                    values['x_compte_comptable'] = str(num_compta)

                if not self.dry_run:
                    domain = ([('ref', '=', values['ref'])] if values.get('ref')
                              else [('name', '=', values['name'])])
                    existing = Partner.search(domain, limit=1)
                    if existing:
                        existing.write(values)
                        stats['updated'] += 1
                    else:
                        Partner.create(values)
                        stats['created'] += 1
                else:
                    stats['created'] += 1

            except Exception as e:
                stats['errors'] += 1
                _logger.error("Import client ligne %s: %s", i + 2, e)

        wb.close()
        return stats

    # =========================================================================
    # IMPORT PRODUITS
    # =========================================================================
    def _import_products(self):
        wb = self._load_workbook()
        sheet_name = 'Trame articles'
        if sheet_name not in wb.sheetnames:
            raise UserError(_(f"Feuille '{sheet_name}' non trouvée dans le fichier."))

        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        data_rows = rows[1:]

        if self.import_limit > 0:
            data_rows = data_rows[:self.import_limit]

        ProductTemplate = self.env['product.template']
        stats = {'total': len(data_rows), 'created': 0, 'updated': 0, 'skipped': 0, 'errors': 0}
        category_cache = {}
        tax_cache = {}

        for i, row in enumerate(data_rows):
            try:
                designation = self._get_cell(row, 4)
                if not designation:
                    stats['skipped'] += 1
                    continue

                code_puit = self._get_cell(row, 0)
                values = {
                    'name': designation,
                    'default_code': str(code_puit) if code_puit else False,
                    'type': 'product',
                    'sale_ok': True,
                    'purchase_ok': True,
                    'tracking': 'lot',
                }

                ean13 = self._get_cell(row, 3)
                if ean13:
                    values['barcode'] = str(ean13)

                pv_ht = self._get_float(row, 19)
                if pv_ht > 0:
                    values['list_price'] = pv_ht
                pr_ht = self._get_float(row, 21)
                if pr_ht > 0:
                    values['standard_price'] = pr_ht
                poids_net = self._get_float(row, 12)
                if poids_net > 0:
                    values['weight'] = poids_net
                volume = self._get_float(row, 14)
                if volume > 0:
                    values['volume'] = volume

                rayon = self._get_cell(row, 64) or 'Divers'
                groupe = self._get_cell(row, 66)
                sous_groupe = self._get_cell(row, 68)

                rayon_key = f"r:{rayon}"
                if rayon_key not in category_cache:
                    category_cache[rayon_key] = self._find_or_create_product_category(rayon)

                categ_id = category_cache[rayon_key]
                if groupe:
                    grp_key = f"g:{rayon}:{groupe}"
                    if grp_key not in category_cache:
                        category_cache[grp_key] = self._find_or_create_product_category(
                            groupe, categ_id)
                    categ_id = category_cache[grp_key]
                    if sous_groupe:
                        sg_key = f"sg:{rayon}:{groupe}:{sous_groupe}"
                        if sg_key not in category_cache:
                            category_cache[sg_key] = self._find_or_create_product_category(
                                sous_groupe, categ_id)
                        categ_id = category_cache[sg_key]

                values['categ_id'] = categ_id

                taux_tva = self._get_cell(row, 18)
                if taux_tva is not None:
                    try:
                        taux = float(taux_tva)
                        tax_name = TVA_MAPPING.get(taux)
                        if tax_name:
                            if tax_name not in tax_cache:
                                tax = self.env['account.tax'].search(
                                    [('name', 'ilike', tax_name)], limit=1)
                                tax_cache[tax_name] = tax.id if tax else False
                            if tax_cache[tax_name]:
                                values['taxes_id'] = [(6, 0, [tax_cache[tax_name]])]
                    except (ValueError, TypeError):
                        pass

                custom = {
                    'x_code_interne': self._get_cell(row, 1),
                    'x_nom_long': self._get_cell(row, 5),
                    'x_contenu': self._get_float(row, 7),
                    'x_conditionnement': self._get_float(row, 8),
                    'x_pcb': self._get_float(row, 9),
                    'x_poids_brut': self._get_float(row, 11),
                    'x_coef_approche': self._get_float(row, 34, 1.0),
                    'x_zone_peche': self._get_cell(row, 48),
                    'x_nom_scientifique': self._get_cell(row, 49),
                    'x_origine': self._get_cell(row, 62),
                    'x_marque': self._get_cell(row, 61),
                    'x_segment': self._get_cell(row, 70),
                    'x_taux_om': self._get_float(row, 56),
                    'x_pv_ttc': self._get_float(row, 20),
                    'x_pr_ttc': self._get_float(row, 22),
                    'x_prmup': self._get_float(row, 23),
                    'x_tarif_t1_ht': self._get_float(row, 35),
                    'x_tarif_t2_ht': self._get_float(row, 36),
                    'x_tarif_t3_ht': self._get_float(row, 37),
                    'x_tarif_t4_ht': self._get_float(row, 38),
                    'x_tarif_t5_ht': self._get_float(row, 39),
                    'x_tarif_t6_ht': self._get_float(row, 40),
                }
                for k, v in custom.items():
                    if v is not None and v != '' and v != 0:
                        values[k] = v if not isinstance(v, (int, float)) or v != 0 else v

                if not self.dry_run:
                    domain = ([('default_code', '=', values['default_code'])]
                              if values.get('default_code')
                              else [('name', '=', values['name'])])
                    existing = ProductTemplate.search(domain, limit=1)
                    if existing:
                        existing.write(values)
                        stats['updated'] += 1
                    else:
                        ProductTemplate.create(values)
                        stats['created'] += 1
                else:
                    stats['created'] += 1

            except Exception as e:
                stats['errors'] += 1
                _logger.error("Import produit ligne %s: %s", i + 2, e)

        wb.close()
        return stats

    # =========================================================================
    # ACTION PRINCIPALE
    # =========================================================================
    def action_import(self):
        self.ensure_one()

        if self.import_type == 'suppliers':
            stats = self._import_suppliers()
            title = 'FOURNISSEURS'
        elif self.import_type == 'clients':
            stats = self._import_clients()
            title = 'CLIENTS'
        elif self.import_type == 'products':
            stats = self._import_products()
            title = 'PRODUITS'
        else:
            raise UserError(_("Type d'import non reconnu."))

        mode = "SIMULATION" if self.dry_run else "RÉEL"
        self.result_text = (
            f"{'=' * 50}\n"
            f"RÉSULTATS IMPORT {title} (MODE {mode})\n"
            f"{'=' * 50}\n"
            f"  Total lignes     : {stats['total']}\n"
            f"  Créés            : {stats['created']}\n"
            f"  Mis à jour       : {stats['updated']}\n"
            f"  Ignorés          : {stats['skipped']}\n"
            f"  Erreurs          : {stats['errors']}\n"
            f"{'=' * 50}\n"
        )
        self.state = 'done'

        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }

    def action_reset(self):
        self.state = 'draft'
        self.result_text = False
        return {
            'type': 'ir.actions.act_window',
            'res_model': self._name,
            'res_id': self.id,
            'view_mode': 'form',
            'target': 'new',
        }
